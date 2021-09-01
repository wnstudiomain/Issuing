from openpyxl import load_workbook
from os import path


class ParseXLSX:
    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def search_value_in_row(self, value):
        wb = load_workbook(self.file)  # Load the workbook
        ws = wb[self.sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell == value:
                    return [x for x in row]
